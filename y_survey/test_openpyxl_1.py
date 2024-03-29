'''
Created on 2019年9月10日

@author: 静火
'''
import openpyxl
 
workbook = openpyxl.load_workbook(r"2018.xlsx")
ws = workbook.active
print (ws.title)
# 读取一行或者一列
for cell in ws['1']:  # 行
# for cell in ws['A']:#列
    if cell.coordinate == 'BB1':
        break
    print(cell.coordinate, cell.value)  # coordinate 坐标
# worksheets=workbook.get_sheet_names()# 获取所有sheet名称,方法已废弃
worksheets = workbook.sheetnames  # 方法当属性用@property
 
# sheet=workbook.get_sheet_by_name(worksheets[1])#获取/激活一个工作表,方法已废弃
sheet = workbook[worksheets[1]]  # 获取/激活一个工作表
 
# 两种读取方式
print(sheet['B1'], sheet['B1'].value)
print(sheet.cell(row=1, column=2).value)
# 获取最大行数
print(sheet.max_row, sheet.max_column)
 
# 查询统计表的名称
for r in range(1, sheet.max_row - 1):
    name = sheet.cell(row=r, column=2).value
    value = sheet.cell(row=r, column=sheet.max_column - 1).value
    if not isinstance(value, int) :  # 判断类型
        print('pass')
    print(name, value)
# 字母和数字之间转换
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(11))
print(column_index_from_string('A'))
 
# 获取A2:B4 单元格区域
print('************************************')
print(tuple(sheet["A2":"B4"]))
for rowOfCell in sheet["A2":"B4"]:
    for temp in rowOfCell:
        if __name__ == '__main__':
            print(temp.coordinate, temp.value)
