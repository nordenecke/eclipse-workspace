# -*- coding: utf-8 -*-
'''
Created on 2019��9��10��

@author: ����
'''
import openpyxl
 
workbook = openpyxl.load_workbook(r"2018.xlsx")
ws = workbook.active
print (ws.title)
# ��ȡһ�л���һ��
for cell in ws['1']:  # ��
# for cell in ws['A']:#��
    if cell.coordinate == 'BB1':
        break
    print(cell.coordinate, cell.value)  # coordinate ����
# worksheets=workbook.get_sheet_names()# ��ȡ����sheet����,�����ѷ���
worksheets = workbook.sheetnames  # ������������@property
 
# sheet=workbook.get_sheet_by_name(worksheets[1])#��ȡ/����һ��������,�����ѷ���
sheet = workbook[worksheets[1]]  # ��ȡ/����һ��������
 
# ���ֶ�ȡ��ʽ
print(sheet['B1'], sheet['B1'].value)
print(sheet.cell(row=1, column=2).value)
# ��ȡ�������
print(sheet.max_row, sheet.max_column)
 
# ��ѯͳ�Ʊ������
for r in range(1, sheet.max_row - 1):
    name = sheet.cell(row=r, column=2).value
    value = sheet.cell(row=r, column=sheet.max_column - 1).value
    if not isinstance(value, int) :  # �ж�����
        print('pass')
    print(name, value)
# ��ĸ������֮��ת��
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(11))
print(column_index_from_string('A'))
 
# ��ȡA2:B4 ��Ԫ������
print('************************************')
print(tuple(sheet["A2":"B4"]))
for rowOfCell in sheet["A2":"B4"]:
    for temp in rowOfCell:
        if __name__ == '__main__':
            print(temp.coordinate, temp.value)
