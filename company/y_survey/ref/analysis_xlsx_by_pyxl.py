'''
Created on 2019��9��10��

@author: ����
'''

# coding:utf-8
import xlrd
import xlwt
# ��д2007 excel
import openpyxl
import sys


# ��ȡ�豸sn
# def readSN(path):
#   wb = openpyxl.load_workbook(path)
#   sheet = wb.active
#   dict = []
#   for i in range(2, sheet.max_row +1):
#     c = sheet["C" + str(i)].value;
#     d = sheet["D" + str(i)].value;
#
#     dict.append(d)
#     #dict.append(d)
#     #print(c,d)
#   return dict;
#
#   pass;
# print(readSN("./sim/1.xlsx"))
def read07Excel(path, path1):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    # print(sheet.max_column) # ��ȡ�������
    # print(sheet.max_row) # ��ȡ�������
    # print(sheet['B1'].value)
    wb1 = openpyxl.load_workbook(path1)
    sheet1 = wb1.active
    for i in range(2, sheet.max_row):
        iccid = sheet["B" + str(i)].value;
        len_iccid = len(iccid)
        if len_iccid == 20 :
            sub_iccid = iccid[16:-1]
        elif len_iccid == 21:
            sub_iccid = iccid[17:-1]
        for x in range(1, sheet1.max_row):
            # print(sheet1["D"+str(x)].value)
            if sub_iccid + "N" == sheet1["D" + str(x)].value:
                sheet["O" + str(i)].value = sheet1["C" + str(x)].value;
                wb.save(filename=path)
                print(str(sheet1["D" + str(x)].value) + " " + str(sheet1["C" + str(x)].value) + " " + str(iccid))
                print()
            pass


# д������
# s =sheet["P"+str(i)].value = "dsdaf";
# wb.save(filename=path)
# p = sheet["P" + str(i)].value;
# print(sub_iccid)
# for row in sheet.rows:
#   for cell in row:
#     print(cell.value, "\t", end="")
#     print(cell.column, "\t", end="")
#
#
#   print()
#   sys.exit()
# path = "./sim/2.xlsx"
# wb = openpyxl.load_workbook(path)
# #sheet = wb.sheetnames[0] #��ȡ����
# sheet = wb.active
# �ֱ𷵻�
# print(sheet['A1'].value) #��ȡ��Ԫ��A1ֵ
read07Excel("./sim/2.xlsx", "./sim/1.xlsx")
# wb=openpyxl.load_workbook('./sim/1.xlsx') #��excel�ļ�
# print(wb.sheetnames) #��ȡ���������й�������