'''
Created on 2019��9��10��

@author: ����
'''
from openpyxl import load_workbook  # �����ȡexcel�ļ���ģ��
from openpyxl import Workbook  # �����½�excel�ļ���ģ��

xls_read = load_workbook('pyxl_test.xlsx')  # ��excel�ļ���Ϊ'pyxl_test.xlsx'
# print(xls_read.sheetnames)  # �鿴������'pyxl_test.xlsx'�е�����sheet�������б���ʽ����

print(xls_read.active)  # �鿴�ļ�pyxl_test�Ļ��sheet
# xls_read.active.title = 'test'  # ����е�sheet���Ʊ��Ϊtest
xls_read_sheet = xls_read.active  # ����е�sheet��ֵ������

# xls_read_sheet = xls_read.get_sheet_by_name('test') # ��ȡexcel�ļ���ĳһ��sheet

# print(xls_read_sheet['C'])    # ��ȡsheet�е�C���������ݣ���������Ԫ����ʽ����
# print(xls_read_sheet['C4'].value) # ��ȡsheet��'C4'��Ԫ���ֵ
# print(xls_read_sheet.max_column)  # �鿴sheet�������кϼ�ֵ��ͳ��������ֻҪ��Ԫ����ֵ������һ��
# print(xls_read_sheet.max_row)     # �鿴sheet�������кϼ�ֵ��ͳ��������ֻҪ��Ԫ����ֵ������һ��

# b4 = xls_read_sheet['B4']   # ͨ���к�+���� ����λĳһ��cell
# print(b4.value)   # ʹ��.value ����ȡĳһ��cell��ֵ

# print(xls_read_sheet.cell(column=2,row=4).value)  # ͨ��ĳsheet.cell(column=?,row=?).value ��ȡ��ĳһ����Ԫ���ֵ

# xls_read_sheet.rows # sheet.rows��һ������������ÿһ�е������γ�һ��Ԫ��
# for row in xls_read_sheet.rows:
#     print(row)
#     for cell in row:
#         print(cell.value)
#

# for column in xls_read_sheet.columns:   #sheet.columns��һ��������������ÿһ�� һ�е������γ�Ԫ��
#     print(list(column))
#     print(list(column)[0].value)
# for i in column:
#     if i.value:
#         print(i.value)
# print(i.value)

for row in range(1, xls_read_sheet.max_row + 1):
    for col in range(1, xls_read_sheet.max_column + 1):
        res = xls_read_sheet.cell(row=row, column=col)
        if res.value:
            print(res.value, end=' ')
            # print(xls_read_sheet.cell(row=row,column=col).value,end=' ')
    print()

xls_read.save('pyxl_test.xlsx')  # ������ļ�

print('=' * 40)

wb = Workbook()  # �½���һ�Ź�������Ĭ�ϴ�����һ������'Sheet'��sheet��
print(wb)  # <openpyxl.workbook.workbook.Workbook object at 0x10515bc50>
print(wb.get_sheet_names())  # ��ʾwb�����������е�sheet���õ�һ���б�

wb.create_sheet('Data', index=1)  # ��wb���������½�һ������'Data'��sheet����sheet�������1
# print(wb.get_sheet_names())
# del wb['Sheet']     #ɾ��wb������������'Sheet'��sheet
print(wb.get_sheet_names())
print(wb.active)  # �鿴wb�������л�е�sheet
print(wb.active.values)  # ����wb�������л�е�sheet�������γ�һ��������

wb.active.title = 'test_sheet'  # ��ǰ��е�sheet����
# print(wb.sheetnames)
wb.active['A1'] = 4
wb.active['B1'] = 2
wb.active['C1'] = '=AVERAGE(A1:B1)'  # ʹ��excel�Ĺ�ʽ������ͨ��load_workbook data_only=True��ò��Ҳ�ò���ֵ��
wb.active['D1'] = '=A1*B1'

print(wb.active['A1'].value)
print(wb.active['B1'].value)
print(wb.active['C1'].value)
wb.save('pyxl_test1.xlsx')
