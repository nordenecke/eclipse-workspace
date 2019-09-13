#encoding:utf-8

'''
Created on Sep 11, 2019

@author: eqhuliu
'''
from openpyxl import load_workbook

INPUT_FILES_BASE_PATH = ".\\data\\samples\\"
EXCEL_FILENAME = 'test_sample_1.xlsx'

def get_excelTitleList():
    templist = []
    titlelist = []
    for cx in range(ws.min_column,ws.max_column):
        templist.append(ws.cell(row=1, column=cx).value)
    for i in range(len(templist)):
        if templist[i] != None:
            titlelist.append(templist[i])
    # print('titleList is : ',end= '\t')
    # print(titlelist)
    return titlelist

if __name__=='__main__':
    #获取excel工作页面
    wb = load_workbook(r"{}".format(INPUT_FILES_BASE_PATH) + "{}".format(EXCEL_FILENAME))
    #获取sheet的名称
    sheetNames = wb.get_sheet_names()
    print(sheetNames) 
    #获取sheet0中的表头
    ws = wb.get_sheet_by_name(sheetNames[0])
    print(get_excelTitleList())
